"""Chunked, auditable import for the full performance and customer extracts."""
from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Iterator

from openpyxl import load_workbook

from services.customer_fact_refresh import policy_identity_sql, policy_key_sql, refresh_customer_facts


PERFORMANCE_COLUMNS = [
    "年", "年季", "年月", "年月日", "销售机构名称", "项目名称", "业务模式", "人员工号",
    "主管工号", "经理工号", "投保单号", "是否企划口径", "证券方营业网点名称",
    "证券方销售人员工号", "自保件标记", "投保时间", "承保时间", "入账时间", "回销时间",
    "犹豫期退保时间", "产品代码", "产品名称", "长短险", "缴费年限", "保障年限", "当前缴别",
    "是否商保年金产品", "是否社会保障型产品", "产品设计分类", "是否个人养老金", "产品类型",
    "投保人id", "期交保费", "折算保费", "年化规保", "价值规保", "承保件数",
]
CUSTOMER_COLUMNS = [
    "投保单号", "投保人id", "投保时间", "导入时间", "回销时间", "承保时间", "入账时间",
    "犹豫期退保时间", "保单状态名称", "保单终止原因",
]
NUMERIC_PERFORMANCE_COLUMNS = {"年", "期交保费", "折算保费", "年化规保", "价值规保", "承保件数"}
TRANSFORM_LINE_SQL = "CASE TRIM(\"业务模式\") WHEN '证券' THEN '证保' WHEN '网服' THEN '蚁桥' ELSE TRIM(\"业务模式\") END"


@dataclass(frozen=True)
class ImportResult:
    batch_id: int
    performance_rows: int
    customer_source_rows: int
    customer_policy_rows: int
    fact_rows: int
    source_text_issue_rows: int
    source_cutoff: str | None
    quick_check: str


def _quote(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value) -> float | int | None:
    text = _text(value).replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def _normalize_source_text(value: str) -> tuple[str, int]:
    if not any(0xDC80 <= ord(char) <= 0xDCFF for char in value):
        return value, int("�" in value)
    raw = value.encode("utf-8", "surrogateescape")
    for encoding in ("gb18030", "latin-1"):
        try:
            return raw.decode(encoding), 1
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", "replace"), 1


def _find_excel_header(sheet) -> tuple[int, list[str]]:
    required = {"投保单号", "期交保费", "业务模式"}
    for row_no, values in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        columns = [_text(value) for value in values]
        if required.issubset(set(columns)):
            return row_no, columns
    raise ValueError(f"{sheet.title} 前20行未找到业绩表头")


class FullHistoryImporter:
    """Build full history in an offline SQLite copy, then validate it for atomic promotion."""

    def __init__(self, database: str | Path, source_directory: str | Path, *, imported_by: str = "system"):
        self.database = Path(database).resolve()
        self.source_directory = Path(source_directory).resolve()
        self.imported_by = imported_by
        self.conn = sqlite3.connect(self.database)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys=ON")
        self.conn.execute("PRAGMA busy_timeout=60000")
        self.conn.execute("PRAGMA temp_store=FILE")
        self.conn.execute("PRAGMA cache_size=-500000")

    def close(self) -> None:
        self.conn.close()

    def _files(self) -> tuple[list[Path], list[Path]]:
        performance = sorted(self.source_directory.glob("AI-电商业绩_*.csv"))
        performance += sorted(self.source_directory.glob("AI-电商业绩_*.xlsx"))
        customers = sorted(self.source_directory.glob("AI-客户清单_*.csv"))
        if len(performance) != 12 or len(customers) != 5:
            raise ValueError(f"全量目录文件不完整：业绩{len(performance)}份、客户{len(customers)}份，预期12份和5份")
        return performance, customers

    def _create_batch(self) -> int:
        cur = self.conn.execute(
            "INSERT INTO history_import_batches (source_directory, imported_by) VALUES (?, ?)",
            (str(self.source_directory), self.imported_by),
        )
        self.conn.commit()
        return int(cur.lastrowid)

    def _create_performance_stage(self) -> None:
        self.conn.execute("DROP TABLE IF EXISTS performance_full_stage")
        definitions = []
        for column in PERFORMANCE_COLUMNS:
            kind = "REAL" if column in NUMERIC_PERFORMANCE_COLUMNS else "TEXT"
            definitions.append(f"{_quote(column)} {kind}")
        self.conn.execute(f"CREATE TABLE performance_full_stage ({', '.join(definitions)})")
        self.conn.commit()

    def _performance_rows_csv(self, path: Path) -> Iterator[tuple[list, int]]:
        with path.open("r", encoding="utf-8-sig", errors="surrogateescape", newline="") as handle:
            reader = csv.DictReader(handle)
            columns = [str(value or "").strip() for value in (reader.fieldnames or [])]
            if columns != PERFORMANCE_COLUMNS:
                raise ValueError(f"{path.name} 字段不一致：预期37列，实际{len(columns)}列")
            for row in reader:
                values = []
                row_has_text_issue = 0
                for column in PERFORMANCE_COLUMNS:
                    value, repaired = _normalize_source_text(row.get(column, ""))
                    row_has_text_issue = max(row_has_text_issue, repaired)
                    values.append(_number(value) if column in NUMERIC_PERFORMANCE_COLUMNS else _text(value))
                yield values, row_has_text_issue

    def _performance_rows_xlsx(self, path: Path) -> Iterator[tuple[list, int]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            header_row, columns = _find_excel_header(sheet)
            index = {column: position for position, column in enumerate(columns) if column}
            missing = [column for column in PERFORMANCE_COLUMNS if column not in index]
            if missing:
                raise ValueError(f"{path.name} 缺少字段：{missing}")
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(value is not None and _text(value) for value in row):
                    continue
                values = []
                for column in PERFORMANCE_COLUMNS:
                    value = row[index[column]] if index[column] < len(row) else None
                    values.append(_number(value) if column in NUMERIC_PERFORMANCE_COLUMNS else _text(value))
                yield values, 0
        finally:
            workbook.close()

    def _record_file(self, batch_id: int, kind: str, path: Path, row_count: int, min_period: str | None,
                     max_period: str | None, text_issues: int) -> None:
        self.conn.execute(
            """INSERT INTO history_import_files
               (batch_id, source_kind, file_name, file_hash, file_size, row_count, min_period, max_period, text_issue_count)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (batch_id, kind, path.name, _hash_file(path), path.stat().st_size, row_count, min_period, max_period, text_issues),
        )

    def _load_performance(self, batch_id: int, paths: Iterable[Path]) -> tuple[int, int]:
        self._create_performance_stage()
        placeholders = ",".join("?" for _ in PERFORMANCE_COLUMNS)
        insert_sql = f"INSERT INTO performance_full_stage VALUES ({placeholders})"
        total = repairs_total = 0
        for path in paths:
            iterator = self._performance_rows_xlsx(path) if path.suffix.lower() == ".xlsx" else self._performance_rows_csv(path)
            batch: list[list] = []
            rows = repairs = 0
            min_period = max_period = None
            for values, repaired in iterator:
                period = _text(values[PERFORMANCE_COLUMNS.index("年月")])[:7]
                if period:
                    min_period = min(min_period or period, period)
                    max_period = max(max_period or period, period)
                batch.append(values)
                repairs += repaired
                rows += 1
                if len(batch) >= 10_000:
                    self.conn.executemany(insert_sql, batch)
                    batch.clear()
            if batch:
                self.conn.executemany(insert_sql, batch)
            self._record_file(batch_id, "performance", path, rows, min_period, max_period, repairs)
            self.conn.commit()
            total += rows
            repairs_total += repairs
        self.conn.execute('CREATE INDEX ix_perf_stage_period_line ON performance_full_stage("年月", "业务模式")')
        self.conn.execute('CREATE INDEX ix_perf_stage_policy ON performance_full_stage("投保单号")')
        self.conn.commit()
        return total, repairs_total

    def _reconcile(self, batch_id: int) -> dict:
        # Reconciliation must retain the same exact-only identities as customer facts.
        # Stage collisions are included before the stage becomes the authoritative raw table.
        self.conn.execute("DROP TABLE IF EXISTS temp.customer_reconcile_ambiguities")
        self.conn.execute("CREATE TEMP TABLE customer_reconcile_ambiguities(policy_key TEXT PRIMARY KEY)")
        self.conn.execute("INSERT INTO customer_reconcile_ambiguities SELECT policy_key FROM customer_policy_key_ambiguity")
        for table in ("performance", "performance_full_stage"):
            columns = {row[1] for row in self.conn.execute(f'PRAGMA table_info("{table}")')}
            if "投保单号" in columns:
                key = policy_key_sql('"投保单号"')
                self.conn.execute(f"""INSERT OR IGNORE INTO customer_reconcile_ambiguities
                    SELECT {key} FROM {table} WHERE {key} IS NOT NULL
                    GROUP BY {key} HAVING COUNT(DISTINCT TRIM(CAST("投保单号" AS TEXT)))>1""")
        def identity(expression):
            return policy_identity_sql(expression, ambiguity_table="temp.customer_reconcile_ambiguities")
        raw_key_expression = identity('"投保单号"')
        self.conn.execute("DROP TABLE IF EXISTS temp.existing_policy_keys")
        self.conn.execute("DROP TABLE IF EXISTS temp.source_policy_keys")
        existing_columns = {row[1] for row in self.conn.execute('PRAGMA table_info("performance")')}
        if "投保单号" in existing_columns:
            self.conn.execute(
                f"""CREATE TEMP TABLE existing_policy_keys AS
                    SELECT DISTINCT {identity('"投保单号"')} AS policy_key,
                           substr(CAST("年月" AS TEXT), 1, 7) AS period,
                           {TRANSFORM_LINE_SQL} AS business_line
                    FROM performance
                    WHERE TRIM(COALESCE("投保单号", '')) <> ''"""
            )
        else:
            self.conn.execute(
                "CREATE TEMP TABLE existing_policy_keys (policy_key TEXT, period TEXT, business_line TEXT)"
            )
        self.conn.execute("CREATE INDEX temp.ix_existing_policy_keys ON existing_policy_keys(period, business_line, policy_key)")
        self.conn.execute(
            f"""CREATE TEMP TABLE source_policy_keys AS
                SELECT DISTINCT {identity('s."投保单号"')} AS policy_key,
                       substr(CAST(s."年月" AS TEXT), 1, 7) AS period,
                       {TRANSFORM_LINE_SQL.replace('"业务模式"', 's."业务模式"')} AS business_line
                FROM performance_full_stage s
                WHERE TRIM(COALESCE(s."投保单号", '')) <> ''
                  AND substr(CAST(s."年月" AS TEXT), 1, 7) IN
                      (SELECT DISTINCT period FROM existing_policy_keys)"""
        )
        self.conn.execute("CREATE INDEX temp.ix_source_policy_keys ON source_policy_keys(period, business_line, policy_key)")
        existing_policy_count = (
            f'COUNT(DISTINCT {raw_key_expression})' if "投保单号" in existing_columns else "0"
        )
        existing = {
            (row[0], row[1]): row[2:]
            for row in self.conn.execute(
                f"""SELECT substr(CAST("年月" AS TEXT),1,7), {TRANSFORM_LINE_SQL}, COUNT(*),
                           SUM(COALESCE("期交保费",0)), {existing_policy_count}
                    FROM performance GROUP BY 1,2"""
            )
        }
        source = {
            (row[0], row[1]): row[2:]
            for row in self.conn.execute(
                f"""SELECT substr(CAST("年月" AS TEXT),1,7), {TRANSFORM_LINE_SQL}, COUNT(*),
                           SUM(COALESCE("期交保费",0)), COUNT(DISTINCT {identity('"投保单号"')})
                    FROM performance_full_stage GROUP BY 1,2"""
            )
        }
        matched_by_key = {
            (row[0], row[1]): int(row[2])
            for row in self.conn.execute(
                """SELECT s.period, s.business_line, COUNT(*)
                   FROM source_policy_keys s
                   JOIN existing_policy_keys e
                     ON e.period=s.period AND e.business_line=s.business_line AND e.policy_key=s.policy_key
                   GROUP BY s.period, s.business_line"""
            )
        }
        rows = []
        for key in sorted(set(existing) | set(source)):
            period, line = key
            old_rows, old_qj, old_policies = existing.get(key, (0, 0, 0))
            new_rows, new_qj, new_policies = source.get(key, (0, 0, 0))
            matched = matched_by_key.get(key, 0)
            decision = "源文件更新替换" if old_rows else "历史补入"
            rows.append((batch_id, period, line, old_rows, new_rows, old_qj or 0, new_qj or 0,
                         old_policies, new_policies, matched, decision))
        self.conn.executemany(
            """INSERT INTO history_reconciliation
               (batch_id, period, business_line, existing_rows, source_rows, existing_qj_premium,
                source_qj_premium, existing_policy_count, source_policy_count, matched_policy_count, decision)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            rows,
        )
        overlap = [row for row in rows if row[3] and row[4]]
        result = {
            "periodCount": len({row[1] for row in rows}),
            "overlapPeriodCount": len({row[1] for row in overlap}),
            "existingRows": sum(row[3] for row in rows),
            "sourceRows": sum(row[4] for row in rows),
            "matchedPolicies": sum(row[9] for row in overlap),
        }
        self.conn.commit()
        self.conn.execute("DROP TABLE customer_reconcile_ambiguities")
        return result

    def _activate_performance(self) -> None:
        self.conn.execute("DROP INDEX IF EXISTS ix_raw_performance_ym_line")
        self.conn.execute("DROP INDEX IF EXISTS ix_raw_performance_policy")
        self.conn.execute("ALTER TABLE performance RENAME TO performance_before_full_history")
        self.conn.execute("ALTER TABLE performance_full_stage RENAME TO performance")
        self.conn.execute('CREATE INDEX ix_raw_performance_ym_line ON performance("年月", "业务模式")')
        self.conn.execute('CREATE INDEX ix_raw_performance_policy ON performance("投保单号")')
        self.conn.execute('CREATE INDEX ix_raw_performance_customer ON performance("投保人id")')
        self.conn.execute("DROP TABLE performance_before_full_history")
        self.conn.commit()

    def _create_customer_stage(self) -> None:
        self.conn.execute("DROP TABLE IF EXISTS customer_source_stage")
        columns = ",".join(f"{_quote(column)} TEXT" for column in CUSTOMER_COLUMNS)
        self.conn.execute(f"CREATE TABLE customer_source_stage ({columns})")
        self.conn.commit()

    def _load_customers(self, batch_id: int, paths: Iterable[Path]) -> tuple[int, str | None]:
        self._create_customer_stage()
        insert_sql = f"INSERT INTO customer_source_stage VALUES ({','.join('?' for _ in CUSTOMER_COLUMNS)})"
        total = 0
        source_cutoff = None
        for path in paths:
            rows = 0
            min_period = max_period = None
            batch: list[list[str]] = []
            with path.open("r", encoding="utf-8-sig", errors="strict", newline="") as handle:
                reader = csv.DictReader(handle)
                columns = [str(value or "").strip() for value in (reader.fieldnames or [])]
                if columns != CUSTOMER_COLUMNS:
                    raise ValueError(f"{path.name} 客户字段不一致")
                for row in reader:
                    values = [_text(row.get(column, "")) for column in CUSTOMER_COLUMNS]
                    underwriting = values[CUSTOMER_COLUMNS.index("承保时间")]
                    imported = values[CUSTOMER_COLUMNS.index("导入时间")]
                    min_period = min(min_period or underwriting, underwriting)
                    max_period = max(max_period or underwriting, underwriting)
                    source_cutoff = max(source_cutoff or imported, imported)
                    batch.append(values)
                    rows += 1
                    if len(batch) >= 10_000:
                        self.conn.executemany(insert_sql, batch)
                        batch.clear()
                if batch:
                    self.conn.executemany(insert_sql, batch)
            self._record_file(batch_id, "customer", path, rows, min_period, max_period, 0)
            self.conn.commit()
            total += rows
        self.conn.execute('CREATE INDEX ix_customer_stage_policy ON customer_source_stage("投保单号")')
        self.conn.execute('CREATE INDEX ix_customer_stage_customer ON customer_source_stage("投保人id")')
        self.conn.commit()
        return total, source_cutoff

    def _build_customer_domains(self, batch_id: int) -> tuple[int, int]:
        customer_key = policy_key_sql('"投保单号"')
        ambiguous_source = self.conn.execute(
            f"""SELECT 1 FROM customer_source_stage GROUP BY {customer_key}
                HAVING {customer_key} IS NULL
                    OR COUNT(DISTINCT TRIM(CAST("投保单号" AS TEXT)))>1
                    OR COUNT(DISTINCT TRIM(COALESCE("投保人id",'')))>1
                    OR MIN(TRIM(COALESCE("投保人id",'')))='' LIMIT 1"""
        ).fetchone()
        if ambiguous_source:
            raise ValueError("新客户源存在模糊保单编号或客户归属缺失/冲突，已停止客户域更新，请核对源文件")
        for table in ("customer_policy_month_fact", "customer_master", "customer_policy_snapshot"):
            self.conn.execute(f"DELETE FROM {table}")
        status_group = """CASE
            WHEN MAX(TRIM("保单状态名称"))='有效' THEN 'active'
            WHEN MAX(TRIM("保单状态名称"))='停效' THEN 'suspended'
            WHEN MAX(TRIM(COALESCE("保单终止原因",'')))='退保终止' THEN 'surrender'
            WHEN MAX(TRIM(COALESCE("保单终止原因",'')))='契撤终止'
              OR MAX(TRIM(COALESCE("犹豫期退保时间",'')))<>'' THEN 'cooling_off'
            WHEN MAX(TRIM(COALESCE("保单终止原因",''))) IN ('到期终止','满期终止') THEN 'maturity'
            WHEN MAX(TRIM(COALESCE("保单终止原因",'')))='一年期险种逾期未付终止' THEN 'short_expiry'
            WHEN MAX(TRIM(COALESCE("保单终止原因",''))) IN
                 ('一般理赔终止','死亡理赔终止','理赔解约终止','死亡终止','短期健康险被保人死亡终止') THEN 'claim'
            WHEN MAX(TRIM("保单状态名称"))='终止' THEN 'other_terminated'
            ELSE 'unknown' END"""
        self.conn.execute(
            f"""INSERT INTO customer_policy_snapshot
                (policy_no, customer_id, application_time, import_time, callback_time, underwriting_time,
                 first_account_time, latest_account_time, hesitation_surrender_time, policy_status,
                 termination_reason, status_group, raw_row_count, batch_id)
                SELECT TRIM("投保单号"), MAX(TRIM("投保人id")), MAX("投保时间"), MAX("导入时间"),
                       MAX("回销时间"), MAX("承保时间"), MIN(NULLIF(TRIM("入账时间"),'')),
                       MAX(NULLIF(TRIM("入账时间"),'')), MAX("犹豫期退保时间"),
                       MAX(TRIM("保单状态名称")), MAX(TRIM(COALESCE("保单终止原因",''))),
                       {status_group}, COUNT(*), ?
                FROM customer_source_stage
                GROUP BY TRIM("投保单号")""",
            (batch_id,),
        )
        self.conn.execute(
            """INSERT INTO customer_master
                (customer_id, first_underwriting_time, first_policy_no, total_policy_count,
                 active_policy_count, suspended_policy_count, terminated_policy_count, batch_id)
                SELECT customer_id, MIN(underwriting_time),
                       MIN(CASE WHEN underwriting_time=first_time THEN policy_no END),
                       COUNT(*), SUM(policy_status='有效'), SUM(policy_status='停效'), SUM(policy_status='终止'), ?
                FROM (
                    SELECT s.*, MIN(underwriting_time) OVER (PARTITION BY customer_id) first_time
                    FROM customer_policy_snapshot s
                ) GROUP BY customer_id""",
            (batch_id,),
        )
        refresh_customer_facts(self.conn, batch_id=batch_id)
        policy_count = self.conn.execute("SELECT COUNT(*) FROM customer_policy_snapshot").fetchone()[0]
        fact_count = self.conn.execute("SELECT COUNT(*) FROM customer_policy_month_fact").fetchone()[0]
        self.conn.execute("DROP TABLE customer_source_stage")
        self.conn.commit()
        return policy_count, fact_count

    def run(self) -> ImportResult:
        performance_files, customer_files = self._files()
        batch_id = self._create_batch()
        try:
            performance_rows, repairs = self._load_performance(batch_id, performance_files)
            reconciliation = self._reconcile(batch_id)
            self._activate_performance()
            customer_source_rows, source_cutoff = self._load_customers(batch_id, customer_files)
            customer_policy_rows, fact_rows = self._build_customer_domains(batch_id)
            quick = self.conn.execute("PRAGMA quick_check").fetchone()[0]
            if quick != "ok":
                raise RuntimeError(f"SQLite quick_check failed: {quick}")
            self.conn.execute(
                """UPDATE history_import_batches SET source_cutoff=?, performance_rows=?,
                   customer_source_rows=?, customer_policy_rows=?, source_text_issue_rows=?, status='success',
                   reconciliation_json=?, completed_at=CURRENT_TIMESTAMP WHERE id=?""",
                (source_cutoff, performance_rows, customer_source_rows, customer_policy_rows, repairs,
                 json.dumps(reconciliation, ensure_ascii=False), batch_id),
            )
            self.conn.commit()
            return ImportResult(batch_id, performance_rows, customer_source_rows, customer_policy_rows,
                                fact_rows, repairs, source_cutoff, quick)
        except Exception as exc:
            self.conn.execute(
                "UPDATE history_import_batches SET status='failed', error_message=?, completed_at=CURRENT_TIMESTAMP WHERE id=?",
                (str(exc)[:2000], batch_id),
            )
            self.conn.commit()
            raise
