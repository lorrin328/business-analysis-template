from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from db import (
    get_db,
    get_kpi_data,
    get_org_kpi_data,
    get_payment_period_structure,
    get_platform_data,
    get_product_structure,
    get_target_config,
)


TARGET_CATEGORY_NAMES = {
    "qjPremium": "期交保费",
    "value": "价值保费",
    "shangbao": "商保年金",
    "baozhang": "保障类产品",
    "tenYear": "10年期产品",
}

TARGET_METRICS = ["整体", "经代", "转型业务", "OTO", "证保", "蚁桥"]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="1F4E78", bold=True, size=13)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _rate(actual: float | None, target: float | None) -> float | None:
    if actual is None or target in (None, 0):
        return None
    return actual / target


def _target_year(payload: dict | None, category: str, metric: str) -> float | None:
    categories = (payload or {}).get("categories") or {}
    item = (
        categories
        .get(category, {})
        .get("metrics", {})
        .get(metric, {})
    )
    if not isinstance(item, dict):
        return None
    return _safe_float(item.get("year"))


def _org_target_year(payload: dict | None, org: str, business_line: str, category: str) -> float | None:
    org_targets = (payload or {}).get("orgTargets") or {}
    item = (
        org_targets
        .get(f"{org}|{business_line}", {})
        .get(category, {})
    )
    if not isinstance(item, dict):
        return None
    return _safe_float(item.get("year"))


def _format_cutoff(kpi: dict) -> str:
    daily = kpi.get("daily_cutoff") or {}
    if daily.get("use_daily") and daily.get("month") and daily.get("day"):
        return f"{kpi.get('year')}年{daily.get('month')}月{daily.get('day')}日"
    month = kpi.get("month")
    return f"{kpi.get('year')}年{month}月" if month else f"{kpi.get('year')}年"


def _write_table(ws, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    max_col = max(len(headers), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(value, float):
                cell.number_format = "0.00"
            if isinstance(value, int):
                cell.number_format = "#,##0"

    ws.freeze_panes = "A3"
    if headers:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(2, len(rows) + 2)}"
    _fit_columns(ws)


def _fit_columns(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 38))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)


def _sheet(ws, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    _write_table(ws, title, headers, rows or [["暂无数据"] + [""] * (len(headers) - 1)])


def _flatten_targets(payload: dict | None) -> list[list[Any]]:
    if not payload or not payload.get("categories"):
        return [["未配置服务端目标", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]]

    rows: list[list[Any]] = []
    year = payload.get("year")

    def append_row(scope: str, category_key: str, metric: str, values: dict, org: str = "") -> None:
        rows.append([
            year,
            TARGET_CATEGORY_NAMES.get(category_key, category_key),
            scope,
            org,
            metric,
            _safe_float(values.get("year")) or 0,
            *[(_safe_float(v) or 0) for v in (values.get("quarter") or [0, 0, 0, 0])[:4]],
            *[(_safe_float(v) or 0) for v in (values.get("month") or [0] * 12)[:12]],
        ])

    for category_key, category in (payload.get("categories") or {}).items():
        for metric in TARGET_METRICS:
            values = (category.get("metrics") or {}).get(metric)
            if isinstance(values, dict):
                append_row("整体/业务线", category_key, metric, values)

    for org_key, metrics in (payload.get("orgTargets") or {}).items():
        org, business_line = (org_key.split("|", 1) + [""])[:2] if "|" in org_key else (org_key, "")
        for category_key, values in (metrics or {}).items():
            if isinstance(values, dict):
                append_row("机构", category_key, business_line, values, org)
    return rows


def _kpi_rows(kpi: dict, target_payload: dict | None) -> list[list[Any]]:
    qj = kpi.get("qj_premium") or {}
    value = kpi.get("value") or {}
    hr = kpi.get("hr") or {}
    qj_target = _target_year(target_payload, "qjPremium", "整体")
    value_target = _target_year(target_payload, "value", "整体")
    annuity_target = _target_year(target_payload, "shangbao", "整体")
    protection_target = _target_year(target_payload, "baozhang", "整体")
    tenyear_target = _target_year(target_payload, "tenYear", "整体")
    cutoff = _format_cutoff(kpi)

    active = sum((_safe_float(v.get("active")) or 0) for v in hr.values() if isinstance(v, dict))
    avg = sum((_safe_float(v.get("avg")) or 0) for v in hr.values() if isinstance(v, dict))
    activity_rate = _rate(active, avg)

    return [
        ["期交保费达成率", qj.get("total", 0), qj_target, _rate(qj.get("total", 0), qj_target), "经代+OTO+证保+蚁桥", cutoff],
        ["价值达成率", round(sum((_safe_float(v) or 0) for v in value.values()), 2), value_target, _rate(round(sum((_safe_float(v) or 0) for v in value.values()), 2), value_target), "OTO+证保+蚁桥", f"{kpi.get('year')}年{kpi.get('month')}月"],
        ["长险活动率", activity_rate, None, None, "活动人力/月均在职人力", f"{kpi.get('year')}年{kpi.get('month')}月"],
        ["商保年金达成率", kpi.get("annuity_total", 0), annuity_target, _rate(kpi.get("annuity_total", 0), annuity_target), "经代+转型参数打标产品", f"{kpi.get('year')}年{kpi.get('month')}月"],
        ["保障类产品达成率", kpi.get("protection_total", 0), protection_target, _rate(kpi.get("protection_total", 0), protection_target), "经代+转型参数打标产品", f"{kpi.get('year')}年{kpi.get('month')}月"],
        ["10年期产品达成率", kpi.get("tenyear_total", 0), tenyear_target, _rate(kpi.get("tenyear_total", 0), tenyear_target), "转型10年及以上+经代10年及以上", f"{kpi.get('year')}年{kpi.get('month')}月"],
        ["长险期交达成率", kpi.get("longterm_qj", 0), qj_target, _rate(kpi.get("longterm_qj", 0), qj_target), "长险期交，目标沿用期交保费目标", cutoff],
        ["人均保费", None, None, None, "当前看板展示口径为转型业务，不含经代", cutoff],
    ]


def _org_rows(org_data: dict, target_payload: dict | None) -> list[list[Any]]:
    rows: list[list[Any]] = []
    perf = org_data.get("perf") or {}
    value = org_data.get("value") or {}
    for key in sorted(set(perf.keys()) | set(value.keys())):
        org, line = (key.split("|", 1) + [""])[:2] if "|" in key else (key, "")
        year_perf = (perf.get(key) or {}).get("year") or {}
        value_actual = ((value.get(key) or {}).get("year") or 0)
        qj_actual = year_perf.get("qj_premium", 0)
        qj_target = _org_target_year(target_payload, org, line, "qjPremium")
        value_target = _org_target_year(target_payload, org, line, "value")
        ten_target = _org_target_year(target_payload, org, line, "tenYear")
        annuity_target = _org_target_year(target_payload, org, line, "shangbao")
        protection_target = _org_target_year(target_payload, org, line, "baozhang")
        rows.append([
            org,
            line,
            qj_target,
            qj_actual,
            _rate(qj_actual, qj_target),
            value_target,
            value_actual,
            _rate(value_actual, value_target),
            ten_target,
            year_perf.get("product_10year", 0),
            _rate(year_perf.get("product_10year", 0), ten_target),
            annuity_target,
            year_perf.get("product_annuity", 0),
            _rate(year_perf.get("product_annuity", 0), annuity_target),
            protection_target,
            year_perf.get("product_protection", 0),
            _rate(year_perf.get("product_protection", 0), protection_target),
        ])
    return rows


def _platform_rows(platform: dict) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for r in platform.get("performance") or []:
        rows.append([r.get("month"), r.get("channel"), r.get("qj_premium"), r.get("gm_premium"), r.get("zs_premium")])
    for r in platform.get("jingdai") or []:
        rows.append([r.get("month"), "经代", r.get("qj_premium"), r.get("gm_premium"), r.get("zs_premium")])
    return sorted(rows, key=lambda x: (x[0] or 0, str(x[1] or "")))


def _product_rows(product_data: dict, value_key: str) -> list[list[Any]]:
    return [[r.get("name"), r.get("value")] for r in product_data.get(value_key) or []]


def _payment_rows(payment_data: dict) -> list[list[Any]]:
    premium_map = {r.get("name"): r.get("value") for r in payment_data.get("premium") or []}
    count_map = {r.get("name"): r.get("value") for r in payment_data.get("count") or []}
    return [[name, premium_map.get(name), count_map.get(name)] for name in sorted(set(premium_map) | set(count_map))]


def _team_rows(platform: dict) -> list[list[Any]]:
    rows = []
    for r in platform.get("hr") or []:
        start = _safe_float(r.get("start_headcount")) or 0
        end = _safe_float(r.get("end_headcount")) or 0
        avg = (start + end) / 2
        active = _safe_float(r.get("active_headcount")) or 0
        rows.append([r.get("month"), r.get("channel"), start, end, avg, active, _rate(active, avg)])
    return sorted(rows, key=lambda x: (x[0] or 0, str(x[1] or "")))


def _product_config_rows() -> list[list[Any]]:
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT product_code, product_name, business_type, is_annuity, is_protection
            FROM product_config
            ORDER BY COALESCE(business_type, ''), product_code
            """
        ).fetchall()
        return [[r["product_code"], r["product_name"], r["business_type"], r["is_annuity"], r["is_protection"]] for r in rows]


def build_dashboard_export_workbook(year: int) -> bytes:
    kpi = get_kpi_data(year)
    platform = get_platform_data(year)
    org_data = get_org_kpi_data(year)
    target_payload = get_target_config(year)
    product_design = get_product_structure(year, "design_cat")
    product_mix = get_product_structure(year, "product_mix")
    payment = get_payment_period_structure(year)

    wb = Workbook()
    wb.remove(wb.active)

    exported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws = wb.create_sheet("导出说明")
    _sheet(
        ws,
        f"{year}年经营分析看板导出说明",
        ["项目", "内容"],
        [
            ["导出时间", exported_at],
            ["数据年份", year],
            ["数据截止", _format_cutoff(kpi)],
            ["目标来源", "target_config 服务端目标" if target_payload else "未配置服务端目标"],
            ["说明", "本文件导出目标与各模块表格数据，不包含图表。金额单位为万元。"],
        ],
    )

    ws = wb.create_sheet("KPI概览")
    _sheet(ws, f"{year}年KPI概览", ["指标", "实绩", "目标", "达成率", "口径说明", "时间口径"], _kpi_rows(kpi, target_payload))

    ws = wb.create_sheet("目标设置")
    _sheet(ws, f"{year}年目标设置", ["年份", "目标分类", "层级", "机构", "业务线", "年度", "Q1", "Q2", "Q3", "Q4", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"], _flatten_targets(target_payload))

    ws = wb.create_sheet("机构维度")
    _sheet(ws, f"{year}年机构维度", ["机构", "业务模式", "期交目标", "期交达成", "期交达成率", "价值目标", "价值达成", "价值达成率", "10年期目标", "10年期达成", "10年期达成率", "商保年金目标", "商保年金达成", "商保年金达成率", "保障类目标", "保障类达成", "保障类达成率"], _org_rows(org_data, target_payload))

    ws = wb.create_sheet("平台趋势")
    _sheet(ws, f"{year}年平台月度数据", ["月份", "业务线", "期交保费", "规模保费", "折算保费"], _platform_rows(platform))

    ws = wb.create_sheet("产品结构")
    _sheet(ws, f"{year}年产品结构-设计分类", ["分类", "保费"], _product_rows(product_design, "premium"))

    ws = wb.create_sheet("产品明细")
    _sheet(ws, f"{year}年产品结构-产品明细", ["产品", "保费"], _product_rows(product_mix, "premium"))

    ws = wb.create_sheet("交期结构")
    _sheet(ws, f"{year}年交期结构", ["交期分类", "保费", "件数"], _payment_rows(payment))

    ws = wb.create_sheet("队伍分析")
    _sheet(ws, f"{year}年队伍分析", ["月份", "业务线", "月初人力", "月末人力", "月均人力", "活动人力", "活动率"], _team_rows(platform))

    ws = wb.create_sheet("产品参数")
    _sheet(ws, "产品参数配置", ["产品代码", "产品名称", "业务类型", "商保年金", "保障类"], _product_config_rows())

    percent_columns = {
        "KPI概览": {4},
        "机构维度": {5, 8, 11, 14, 17},
        "队伍分析": {7},
    }
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                if cell.column in percent_columns.get(ws.title, set()) and isinstance(cell.value, float):
                    cell.number_format = "0.0%"
        if ws.title == "KPI概览":
            ws["B5"].number_format = "0.0%"

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
