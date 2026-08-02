"""Chunked upload and background, streaming customer snapshot import jobs."""
from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
import sqlite3
from pathlib import Path
from typing import Iterable
from uuid import uuid4

from openpyxl import load_workbook

import db.connection as connection
from customer_analysis.importer import (
    CUSTOMER_IMPORT_COLUMNS,
    CustomerImportError,
    REQUIRED_COLUMNS,
    _find_header,
    _normalize_headers,
    _text,
    _to_policy,
)
from db.connection import get_db
from services.operation_lock import OperationLockError, operation_lock


CHUNK_BYTES = 8 * 1024 * 1024
READY_TTL_HOURS = 24


def _root() -> Path:
    configured = os.getenv("CUSTOMER_IMPORT_STAGING_DIR", "").strip()
    root = Path(configured) if configured else Path(connection.DB_PATH).resolve().parent / "customer-import-staging"
    root = root.resolve()
    root.mkdir(parents=True, exist_ok=True, mode=0o700)
    try:
        root.chmod(0o700)
    except OSError:
        pass
    return root


def _job_dir(upload_id: str) -> Path:
    if not upload_id or any(char not in "0123456789abcdef-" for char in upload_id.lower()):
        raise CustomerImportError("上传任务编号无效")
    root = _root()
    target = (root / upload_id).resolve()
    if target.parent != root:
        raise CustomerImportError("上传任务路径无效")
    return target


def _file_path(upload_id: str, index: int, suffix: str) -> Path:
    return _job_dir(upload_id) / f"{index:06d}{suffix.lower()}"


def _safe_remove_job(upload_id: str) -> None:
    target = _job_dir(upload_id)
    if target.exists():
        shutil.rmtree(target)


def cleanup_expired_imports() -> None:
    with get_db() as conn:
        rows = conn.execute(
            """SELECT upload_id FROM customer_import_batches
               WHERE upload_id IS NOT NULL AND status IN ('uploading','processing','ready','importing','failed','blocked','expired')
                 AND COALESCE(updated_at, imported_at) < datetime('now', ?)""",
            (f"-{READY_TTL_HOURS} hours",),
        ).fetchall()
        upload_ids = [str(row["upload_id"]) for row in rows]
        if upload_ids:
            conn.executemany(
                """UPDATE customer_import_batches SET status=CASE WHEN status IN ('uploading','processing','ready','importing') THEN 'expired' ELSE status END,
                   updated_at=CURRENT_TIMESTAMP WHERE upload_id=?""",
                [(value,) for value in upload_ids],
            )
            conn.commit()
    for upload_id in upload_ids:
        _safe_remove_job(upload_id)


def create_upload_session(files: list[dict], imported_by: str) -> dict:
    cleanup_expired_imports()
    if not files:
        raise CustomerImportError("请选择至少一份客户清单")
    normalized = []
    total_bytes = 0
    for index, item in enumerate(files):
        name = Path(str(item.get("name") or "")).name.strip()
        try:
            size = int(item.get("size"))
        except (TypeError, ValueError) as exc:
            raise CustomerImportError("文件大小无效") from exc
        suffix = Path(name).suffix.lower()
        if not name or suffix not in {".csv", ".xlsx"}:
            raise CustomerImportError("仅支持.csv或.xlsx客户清单")
        if size <= 0:
            raise CustomerImportError(f"{name}为空文件")
        normalized.append({"index": index, "name": name, "size": size, "suffix": suffix})
        total_bytes += size
    upload_id = str(uuid4())
    target = _job_dir(upload_id)
    if total_bytes > shutil.disk_usage(_root()).free:
        raise CustomerImportError("服务器临时磁盘空间不足，无法接收本批文件")
    target.mkdir(mode=0o700)
    with get_db() as conn:
        cursor = conn.execute(
            """INSERT INTO customer_import_batches
               (imported_by, manifest_hash, file_count, total_bytes, status, upload_id, updated_at)
               VALUES (?, '', ?, ?, 'uploading', ?, CURRENT_TIMESTAMP)""",
            (imported_by, len(normalized), total_bytes, upload_id),
        )
        batch_id = int(cursor.lastrowid)
        conn.executemany(
            """INSERT INTO customer_import_files
               (batch_id, file_name, file_hash, file_size, row_count, file_index, received_bytes)
               VALUES (?, ?, '', ?, 0, ?, 0)""",
            [(batch_id, item["name"], item["size"], item["index"]) for item in normalized],
        )
        conn.commit()
    return {"uploadId": upload_id, "batchId": batch_id, "files": normalized, "totalBytes": total_bytes, "chunkBytes": CHUNK_BYTES, "status": "uploading"}


def append_upload_chunk(upload_id: str, file_index: int, offset: int, chunk: bytes, username: str) -> dict:
    if not chunk:
        raise CustomerImportError("上传分片为空")
    if len(chunk) > CHUNK_BYTES:
        raise CustomerImportError(f"单个上传分片不得超过{CHUNK_BYTES // 1024 // 1024}MB")
    with get_db() as conn:
        job = conn.execute(
            "SELECT id, imported_by, status FROM customer_import_batches WHERE upload_id=?", (upload_id,),
        ).fetchone()
        if not job or job["imported_by"] != username:
            raise CustomerImportError("上传任务不存在或不属于当前账号")
        if job["status"] != "uploading":
            raise CustomerImportError("当前任务不接受上传分片")
        file_row = conn.execute(
            """SELECT id, file_name, file_size, received_bytes FROM customer_import_files
               WHERE batch_id=? AND file_index=?""", (job["id"], file_index),
        ).fetchone()
        if not file_row:
            raise CustomerImportError("上传文件编号不存在")
        if offset != int(file_row["received_bytes"]):
            raise CustomerImportError("上传分片顺序不一致，请从服务器已接收位置续传")
        if offset + len(chunk) > int(file_row["file_size"]):
            raise CustomerImportError("上传内容超过声明的文件大小")
        if shutil.disk_usage(_root()).free < len(chunk) + 64 * 1024 * 1024:
            raise CustomerImportError("服务器临时磁盘剩余空间不足，上传已暂停")
        path = _file_path(upload_id, file_index, Path(file_row["file_name"]).suffix)
        current_size = path.stat().st_size if path.exists() else 0
        if current_size != offset:
            raise CustomerImportError("服务器临时文件与任务进度不一致，请重新创建上传任务")
        with path.open("ab") as handle:
            handle.write(chunk)
        received = offset + len(chunk)
        conn.execute(
            "UPDATE customer_import_files SET received_bytes=? WHERE id=?", (received, file_row["id"]),
        )
        conn.execute(
            """UPDATE customer_import_batches SET received_bytes=(SELECT SUM(received_bytes) FROM customer_import_files WHERE batch_id=?),
               updated_at=CURRENT_TIMESTAMP WHERE id=?""", (job["id"], job["id"]),
        )
        conn.commit()
        total_received = conn.execute("SELECT received_bytes FROM customer_import_batches WHERE id=?", (job["id"],)).fetchone()[0]
    return {"uploadId": upload_id, "fileIndex": file_index, "receivedBytes": received, "totalReceivedBytes": int(total_received)}


def start_processing(upload_id: str, username: str) -> dict:
    with get_db() as conn:
        job = conn.execute(
            "SELECT id, imported_by, status, total_bytes, received_bytes FROM customer_import_batches WHERE upload_id=?",
            (upload_id,),
        ).fetchone()
        if not job or job["imported_by"] != username:
            raise CustomerImportError("上传任务不存在或不属于当前账号")
        if int(job["received_bytes"]) != int(job["total_bytes"]):
            raise CustomerImportError("文件尚未完整上传")
        if job["status"] not in {"uploading", "failed"}:
            raise CustomerImportError("当前任务不能启动预检")
        conn.execute(
            "UPDATE customer_import_batches SET status='processing', error_message=NULL, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (job["id"],),
        )
        conn.commit()
    return {"uploadId": upload_id, "status": "processing"}


def _csv_iterator(path: Path) -> tuple[list[str], Iterable[list[str]]]:
    sample = path.open("rb").read(1024 * 1024)
    encoding = None
    for candidate in ("utf-8-sig", "gb18030"):
        try:
            sample.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    if not encoding:
        raise CustomerImportError(f"{path.name}编码无法识别，请使用UTF-8 BOM、UTF-8或GBK")
    handle = path.open("r", encoding=encoding, errors="strict", newline="")
    reader = csv.reader(handle)
    probe = []
    for _ in range(20):
        try:
            probe.append(next(reader))
        except StopIteration:
            break
    header_no, headers = _find_header(probe, path.name)
    def rows():
        try:
            for row in probe[header_no:]:
                yield row
            yield from reader
        finally:
            handle.close()
    return headers, rows()


def _xlsx_iterator(path: Path) -> tuple[list[str], Iterable[tuple]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        probe = list(sheet.iter_rows(min_row=1, max_row=20, values_only=True))
        try:
            header_no, headers = _find_header(probe, path.name)
        except CustomerImportError:
            continue
        def rows(selected_sheet=sheet, start=header_no + 1):
            try:
                yield from selected_sheet.iter_rows(min_row=start, values_only=True)
            finally:
                workbook.close()
        return headers, rows()
    workbook.close()
    raise CustomerImportError(f"{path.name}各工作表前20行均未找到客户清单表头")


STAGE_COLUMNS = [
    "policy_no", "customer_id", "application_time", "import_time", "callback_time", "underwriting_time",
    "first_account_time", "latest_account_time", "hesitation_surrender_time", "policy_status",
    "termination_reason", "status_group", "raw_row_count",
]


def _stage_values(item) -> tuple:
    return tuple(getattr(item, column) for column in STAGE_COLUMNS)


def _create_stage(path: Path) -> sqlite3.Connection:
    if path.exists():
        path.unlink()
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA temp_store=FILE")
    conn.execute(
        """CREATE TABLE policies (
           policy_no TEXT PRIMARY KEY, customer_id TEXT NOT NULL, application_time TEXT, import_time TEXT NOT NULL,
           callback_time TEXT, underwriting_time TEXT NOT NULL, first_account_time TEXT, latest_account_time TEXT,
           hesitation_surrender_time TEXT, policy_status TEXT NOT NULL, termination_reason TEXT,
           status_group TEXT NOT NULL, raw_row_count INTEGER NOT NULL DEFAULT 1,
           conflict INTEGER NOT NULL DEFAULT 0, action TEXT)"""
    )
    return conn


STAGE_UPSERT = """INSERT INTO policies
    (policy_no, customer_id, application_time, import_time, callback_time, underwriting_time,
     first_account_time, latest_account_time, hesitation_surrender_time, policy_status,
     termination_reason, status_group, raw_row_count)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ON CONFLICT(policy_no) DO UPDATE SET
      conflict=MAX(policies.conflict,
        policies.customer_id<>excluded.customer_id OR
        (policies.import_time=excluded.import_time AND (
          policies.underwriting_time<>excluded.underwriting_time OR
          COALESCE(policies.hesitation_surrender_time,'')<>COALESCE(excluded.hesitation_surrender_time,'') OR
          policies.policy_status<>excluded.policy_status OR
          COALESCE(policies.termination_reason,'')<>COALESCE(excluded.termination_reason,'') OR
          policies.status_group<>excluded.status_group))),
      application_time=CASE WHEN excluded.import_time>policies.import_time THEN excluded.application_time ELSE policies.application_time END,
      callback_time=CASE WHEN excluded.import_time>policies.import_time THEN excluded.callback_time ELSE policies.callback_time END,
      underwriting_time=CASE WHEN excluded.import_time>policies.import_time THEN excluded.underwriting_time ELSE policies.underwriting_time END,
      hesitation_surrender_time=CASE WHEN excluded.import_time>policies.import_time THEN excluded.hesitation_surrender_time ELSE policies.hesitation_surrender_time END,
      policy_status=CASE WHEN excluded.import_time>policies.import_time THEN excluded.policy_status ELSE policies.policy_status END,
      termination_reason=CASE WHEN excluded.import_time>policies.import_time THEN excluded.termination_reason ELSE policies.termination_reason END,
      status_group=CASE WHEN excluded.import_time>policies.import_time THEN excluded.status_group ELSE policies.status_group END,
      import_time=MAX(policies.import_time,excluded.import_time),
      first_account_time=CASE WHEN COALESCE(policies.first_account_time,'')='' THEN excluded.first_account_time
                              WHEN COALESCE(excluded.first_account_time,'')='' THEN policies.first_account_time
                              ELSE MIN(policies.first_account_time,excluded.first_account_time) END,
      latest_account_time=CASE WHEN COALESCE(excluded.latest_account_time,'')='' THEN policies.latest_account_time
                               ELSE MAX(COALESCE(policies.latest_account_time,''),excluded.latest_account_time) END,
      raw_row_count=policies.raw_row_count+excluded.raw_row_count"""


def _update_progress(batch_id: int, rows: int) -> None:
    with get_db() as conn:
        conn.execute(
            "UPDATE customer_import_batches SET processed_rows=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (rows, batch_id),
        )
        conn.commit()


def _assess_stage(stage: sqlite3.Connection) -> dict[str, int]:
    stage.execute("ATTACH DATABASE ? AS prod", (str(Path(connection.DB_PATH).resolve()),))
    try:
        stage.execute(
            """UPDATE policies SET action=CASE
              WHEN conflict=1 THEN 'conflict'
              WHEN NOT EXISTS (SELECT 1 FROM prod.customer_policy_snapshot s WHERE s.policy_no=policies.policy_no) THEN 'insert'
              WHEN customer_id<>(SELECT s.customer_id FROM prod.customer_policy_snapshot s WHERE s.policy_no=policies.policy_no) THEN 'conflict'
              WHEN import_time<COALESCE((SELECT s.import_time FROM prod.customer_policy_snapshot s WHERE s.policy_no=policies.policy_no),'') THEN 'skip_older'
              WHEN import_time=COALESCE((SELECT s.import_time FROM prod.customer_policy_snapshot s WHERE s.policy_no=policies.policy_no),'')
               AND customer_id=(SELECT s.customer_id FROM prod.customer_policy_snapshot s WHERE s.policy_no=policies.policy_no)
               AND underwriting_time=(SELECT s.underwriting_time FROM prod.customer_policy_snapshot s WHERE s.policy_no=policies.policy_no)
               AND COALESCE(hesitation_surrender_time,'')=COALESCE((SELECT s.hesitation_surrender_time FROM prod.customer_policy_snapshot s WHERE s.policy_no=policies.policy_no),'')
               AND policy_status=(SELECT s.policy_status FROM prod.customer_policy_snapshot s WHERE s.policy_no=policies.policy_no)
               AND COALESCE(termination_reason,'')=COALESCE((SELECT s.termination_reason FROM prod.customer_policy_snapshot s WHERE s.policy_no=policies.policy_no),'')
               AND status_group=(SELECT s.status_group FROM prod.customer_policy_snapshot s WHERE s.policy_no=policies.policy_no)
              THEN 'unchanged'
              WHEN import_time=COALESCE((SELECT s.import_time FROM prod.customer_policy_snapshot s WHERE s.policy_no=policies.policy_no),'') THEN 'conflict'
              ELSE 'update' END"""
        )
        stage.commit()
    finally:
        stage.execute("DETACH DATABASE prod")
    counts = {row["action"]: int(row["n"]) for row in stage.execute("SELECT action,COUNT(*) n FROM policies GROUP BY action")}
    return {key: counts.get(key, 0) for key in ("insert", "update", "unchanged", "skip_older", "conflict")}


def prepare_customer_import(upload_id: str) -> None:
    target = _job_dir(upload_id)
    stage_path = target / "normalized.db"
    with get_db() as prod:
        job = prod.execute("SELECT * FROM customer_import_batches WHERE upload_id=?", (upload_id,)).fetchone()
        if not job or job["status"] != "processing":
            return
        batch_id = int(job["id"])
        files = prod.execute(
            "SELECT * FROM customer_import_files WHERE batch_id=? ORDER BY file_index", (batch_id,),
        ).fetchall()
    stage = None
    try:
        stage = _create_stage(stage_path)
        source_rows = invalid_rows = 0
        manifest = []
        for file_row in files:
            path = _file_path(upload_id, int(file_row["file_index"]), Path(file_row["file_name"]).suffix)
            digest = hashlib.sha256()
            with path.open("rb") as handle:
                for block in iter(lambda: handle.read(1024 * 1024), b""):
                    digest.update(block)
            headers, rows = _xlsx_iterator(path) if path.suffix.lower() == ".xlsx" else _csv_iterator(path)
            index = {column: pos for pos, column in enumerate(_normalize_headers(headers)) if column}
            missing = sorted(REQUIRED_COLUMNS - set(index))
            if missing:
                raise CustomerImportError(f"{file_row['file_name']}缺少必需字段：{'、'.join(missing)}")
            buffer = []
            file_rows = 0
            for values in rows:
                if not any(_text(value) for value in values):
                    continue
                file_rows += 1
                source_rows += 1
                raw = {column: (values[pos] if pos < len(values) else None) for column, pos in index.items() if column in CUSTOMER_IMPORT_COLUMNS}
                try:
                    item = _to_policy(raw)
                except CustomerImportError:
                    invalid_rows += 1
                    continue
                buffer.append(_stage_values(item))
                if len(buffer) >= 10_000:
                    stage.executemany(STAGE_UPSERT, buffer)
                    stage.commit()
                    buffer.clear()
                    _update_progress(batch_id, source_rows)
            if buffer:
                stage.executemany(STAGE_UPSERT, buffer)
                stage.commit()
            file_hash = digest.hexdigest()
            manifest.append({"fileName": file_row["file_name"], "fileHash": file_hash, "fileSize": int(file_row["file_size"]), "rowCount": file_rows})
            with get_db() as prod:
                prod.execute(
                    "UPDATE customer_import_files SET file_hash=?, row_count=? WHERE id=?",
                    (file_hash, file_rows, file_row["id"]),
                )
                prod.commit()
            path.unlink()
            _update_progress(batch_id, source_rows)
        normalized = int(stage.execute("SELECT COUNT(*) FROM policies").fetchone()[0])
        duplicate_rows = max(0, source_rows - invalid_rows - normalized)
        counts = _assess_stage(stage)
        source_cutoff = stage.execute("SELECT MAX(import_time) FROM policies").fetchone()[0]
        manifest_hash = hashlib.sha256(json.dumps(manifest, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()
        blocked = invalid_rows > 0 or counts["conflict"] > 0
        error_message = None
        if blocked:
            parts = []
            if invalid_rows:
                parts.append(f"{invalid_rows:,}行必需字段为空或日期无效")
            if counts["conflict"]:
                parts.append(f"{counts['conflict']:,}份保单存在归属或同快照状态冲突")
            error_message = "；".join(parts)
        with get_db() as prod:
            prod.execute(
                """UPDATE customer_import_batches SET manifest_hash=?, source_rows=?, processed_rows=?,
                   normalized_policy_rows=?, inserted_policies=?, updated_policies=?, unchanged_policies=?,
                   skipped_older_policies=?, conflict_policies=?, invalid_rows=?, duplicate_rows=?,
                   source_cutoff=?, status=?, error_message=?, updated_at=CURRENT_TIMESTAMP,
                   completed_at=CASE WHEN ?='blocked' THEN CURRENT_TIMESTAMP ELSE completed_at END WHERE id=?""",
                (manifest_hash, source_rows, source_rows, normalized, counts["insert"], counts["update"],
                 counts["unchanged"], counts["skip_older"], counts["conflict"], invalid_rows,
                 duplicate_rows, source_cutoff, "blocked" if blocked else "ready", error_message,
                 "blocked" if blocked else "ready", batch_id),
            )
            prod.commit()
        if blocked:
            stage.close()
            stage = None
            _safe_remove_job(upload_id)
    except Exception as exc:
        if stage is not None:
            stage.close()
        with get_db() as prod:
            prod.execute(
                """UPDATE customer_import_batches SET status='failed', error_message=?,
                   updated_at=CURRENT_TIMESTAMP, completed_at=CURRENT_TIMESTAMP WHERE upload_id=?""",
                (str(exc)[:500], upload_id),
            )
            prod.commit()
        _safe_remove_job(upload_id)
    finally:
        if stage is not None:
            stage.close()


def request_commit(upload_id: str, username: str) -> dict:
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, imported_by, status FROM customer_import_batches WHERE upload_id=?", (upload_id,),
        ).fetchone()
        if not row or row["imported_by"] != username:
            raise CustomerImportError("上传任务不存在或不属于当前账号")
        if row["status"] != "ready":
            raise CustomerImportError("任务尚未完成预检或存在阻断问题")
        conn.execute(
            "UPDATE customer_import_batches SET status='importing', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (row["id"],),
        )
        conn.commit()
    return {"uploadId": upload_id, "batchId": int(row["id"]), "status": "importing"}


def _commit_prepared_customer_import_unlocked(upload_id: str) -> None:
    target = _job_dir(upload_id)
    stage_path = target / "normalized.db"
    with get_db() as conn:
        job = conn.execute("SELECT * FROM customer_import_batches WHERE upload_id=?", (upload_id,)).fetchone()
        if not job or job["status"] != "importing":
            return
        batch_id = int(job["id"])
        base = conn.execute("SELECT id FROM history_import_batches WHERE status='success' ORDER BY id DESC LIMIT 1").fetchone()
        if not base:
            conn.execute("UPDATE customer_import_batches SET status='failed', error_message='尚未建立全量历史客户库' WHERE id=?", (batch_id,))
            conn.commit()
            return
        try:
            stage_check = sqlite3.connect(stage_path)
            stage_check.row_factory = sqlite3.Row
            try:
                refreshed = _assess_stage(stage_check)
            finally:
                stage_check.close()
            if refreshed["conflict"]:
                conn.execute(
                    """UPDATE customer_import_batches SET status='blocked', conflict_policies=?,
                       error_message='确认导入前数据库状态已变化，请重新上传并预检',
                       updated_at=CURRENT_TIMESTAMP, completed_at=CURRENT_TIMESTAMP WHERE id=?""",
                    (refreshed["conflict"], batch_id),
                )
                conn.commit()
                _safe_remove_job(upload_id)
                return
            conn.execute(
                """UPDATE customer_import_batches SET inserted_policies=?, updated_policies=?,
                   unchanged_policies=?, skipped_older_policies=?, conflict_policies=0,
                   updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                (refreshed["insert"], refreshed["update"], refreshed["unchanged"],
                 refreshed["skip_older"], batch_id),
            )
            conn.commit()
            conn.execute("BEGIN IMMEDIATE")
            conn.execute("ATTACH DATABASE ? AS staged", (str(stage_path),))
            conn.execute(
                """INSERT INTO customer_policy_snapshot
                   (policy_no, customer_id, application_time, import_time, callback_time, underwriting_time,
                    first_account_time, latest_account_time, hesitation_surrender_time, policy_status,
                    termination_reason, status_group, raw_row_count, batch_id, customer_import_batch_id)
                   SELECT policy_no, customer_id, application_time, import_time, callback_time, underwriting_time,
                          first_account_time, latest_account_time, hesitation_surrender_time, policy_status,
                          termination_reason, status_group, raw_row_count, ?, ?
                   FROM staged.policies WHERE action='insert'""", (int(base["id"]), batch_id),
            )
            conn.execute(
                """UPDATE customer_policy_snapshot AS s SET
                   application_time=p.application_time, import_time=p.import_time, callback_time=p.callback_time,
                   underwriting_time=p.underwriting_time,
                   first_account_time=CASE WHEN COALESCE(s.first_account_time,'')='' THEN p.first_account_time
                                           WHEN COALESCE(p.first_account_time,'')='' THEN s.first_account_time
                                           ELSE MIN(s.first_account_time,p.first_account_time) END,
                   latest_account_time=CASE WHEN COALESCE(p.latest_account_time,'')='' THEN s.latest_account_time
                                            ELSE MAX(COALESCE(s.latest_account_time,''),p.latest_account_time) END,
                   hesitation_surrender_time=p.hesitation_surrender_time, policy_status=p.policy_status,
                   termination_reason=p.termination_reason, status_group=p.status_group,
                   raw_row_count=s.raw_row_count+p.raw_row_count, customer_import_batch_id=?
                   FROM staged.policies p WHERE s.policy_no=p.policy_no AND p.action='update'""", (batch_id,),
            )
            conn.execute("DROP TABLE IF EXISTS temp.customer_import_affected_policies")
            conn.execute(
                """CREATE TEMP TABLE customer_import_affected_policies AS
                   SELECT policy_no FROM staged.policies WHERE action IN ('insert','update')"""
            )
            conn.execute("CREATE UNIQUE INDEX temp.ix_customer_import_affected_policy ON customer_import_affected_policies(policy_no)")
            conn.execute("DROP TABLE IF EXISTS temp.customer_import_affected_customers")
            conn.execute(
                """CREATE TEMP TABLE customer_import_affected_customers AS
                   SELECT DISTINCT customer_id FROM customer_policy_snapshot
                   WHERE policy_no IN (SELECT policy_no FROM customer_import_affected_policies)"""
            )
            conn.execute("DELETE FROM customer_master WHERE customer_id IN (SELECT customer_id FROM customer_import_affected_customers)")
            conn.execute(
                """INSERT INTO customer_master
                   (customer_id, first_underwriting_time, first_policy_no, total_policy_count,
                    active_policy_count, suspended_policy_count, terminated_policy_count, batch_id)
                   SELECT customer_id, MIN(underwriting_time),
                          MIN(CASE WHEN underwriting_time=first_time THEN policy_no END), COUNT(*),
                          SUM(policy_status='有效'), SUM(policy_status='停效'), SUM(policy_status='终止'), MAX(batch_id)
                   FROM (SELECT s.*, MIN(underwriting_time) OVER (PARTITION BY customer_id) first_time
                         FROM customer_policy_snapshot s
                         WHERE customer_id IN (SELECT customer_id FROM customer_import_affected_customers))
                   GROUP BY customer_id"""
            )
            conn.execute(
                """UPDATE customer_policy_month_fact SET
                   customer_id=(SELECT s.customer_id FROM customer_policy_snapshot s WHERE s.policy_no=customer_policy_month_fact.policy_no),
                   underwriting_time=(SELECT s.underwriting_time FROM customer_policy_snapshot s WHERE s.policy_no=customer_policy_month_fact.policy_no),
                   policy_status=(SELECT s.policy_status FROM customer_policy_snapshot s WHERE s.policy_no=customer_policy_month_fact.policy_no),
                   termination_reason=(SELECT s.termination_reason FROM customer_policy_snapshot s WHERE s.policy_no=customer_policy_month_fact.policy_no),
                   status_group=(SELECT s.status_group FROM customer_policy_snapshot s WHERE s.policy_no=customer_policy_month_fact.policy_no),
                   customer_match=1 WHERE policy_no IN (SELECT policy_no FROM customer_import_affected_policies)"""
            )
            conn.execute(
                """UPDATE customer_policy_month_fact SET
                   first_customer_underwriting_time=(SELECT m.first_underwriting_time FROM customer_master m WHERE m.customer_id=customer_policy_month_fact.customer_id)
                   WHERE customer_id IN (SELECT customer_id FROM customer_import_affected_customers)"""
            )
            linked = int(conn.execute(
                """SELECT COUNT(*) FROM customer_import_affected_policies p
                   WHERE EXISTS (SELECT 1 FROM customer_policy_month_fact f WHERE f.policy_no=p.policy_no)"""
            ).fetchone()[0])
            conn.execute(
                """UPDATE customer_import_batches SET linked_performance_policies=?, status='success',
                   updated_at=CURRENT_TIMESTAMP, completed_at=CURRENT_TIMESTAMP WHERE id=?""", (linked, batch_id),
            )
            conn.commit()
        except Exception as exc:
            conn.rollback()
            conn.execute(
                """UPDATE customer_import_batches SET status='failed', error_message=?, updated_at=CURRENT_TIMESTAMP,
                   completed_at=CURRENT_TIMESTAMP WHERE id=?""", (str(exc)[:500], batch_id),
            )
            conn.commit()
        finally:
            try:
                conn.execute("DETACH DATABASE staged")
            except sqlite3.Error:
                pass
    _safe_remove_job(upload_id)


def commit_prepared_customer_import(upload_id: str) -> None:
    try:
        with operation_lock("customer-import-background", timeout=60.0):
            _commit_prepared_customer_import_unlocked(upload_id)
    except OperationLockError:
        with get_db() as conn:
            conn.execute(
                """UPDATE customer_import_batches SET status='ready',
                   error_message='当前有其他数据写入任务，请稍后再次确认导入', updated_at=CURRENT_TIMESTAMP
                   WHERE upload_id=? AND status='importing'""", (upload_id,),
            )
            conn.commit()


def get_import_job(upload_id: str, username: str, *, allow_admin: bool = False) -> dict:
    with get_db() as conn:
        row = conn.execute("SELECT * FROM customer_import_batches WHERE upload_id=?", (upload_id,)).fetchone()
        if not row or (row["imported_by"] != username and not allow_admin):
            raise CustomerImportError("上传任务不存在或不属于当前账号")
        files = conn.execute(
            """SELECT file_index, file_name, file_size, received_bytes, row_count
               FROM customer_import_files WHERE batch_id=? ORDER BY file_index""", (row["id"],),
        ).fetchall()
    total = int(row["total_bytes"] or 0)
    received = int(row["received_bytes"] or 0)
    status = str(row["status"])
    return {
        "uploadId": upload_id, "batchId": int(row["id"]), "status": status,
        "fileCount": int(row["file_count"]), "totalBytes": total, "receivedBytes": received,
        "uploadProgress": received / total if total else None,
        "sourceRows": int(row["source_rows"]), "processedRows": int(row["processed_rows"]),
        "normalizedPolicies": int(row["normalized_policy_rows"]),
        "insertPolicies": int(row["inserted_policies"]), "updatePolicies": int(row["updated_policies"]),
        "unchangedPolicies": int(row["unchanged_policies"]), "skippedOlderPolicies": int(row["skipped_older_policies"]),
        "conflictPolicies": int(row["conflict_policies"]), "invalidRows": int(row["invalid_rows"]),
        "duplicateRows": int(row["duplicate_rows"]), "linkedPerformancePolicies": int(row["linked_performance_policies"]),
        "sourceCutoff": row["source_cutoff"], "error": row["error_message"],
        "canImport": status == "ready", "importedAt": row["imported_at"], "completedAt": row["completed_at"],
        "files": [{"index": int(item["file_index"]), "name": item["file_name"], "size": int(item["file_size"]),
                   "receivedBytes": int(item["received_bytes"]), "rowCount": int(item["row_count"])} for item in files],
    }
