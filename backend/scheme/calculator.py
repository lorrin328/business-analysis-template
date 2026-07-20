from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .config import ENTRY_START, MONTHS, RULE_DEFINITIONS, RULE_VERSION, SCHEME_ID, SCHEME_NAME


@dataclass(frozen=True)
class SectionLayout:
    sheet: str
    section: str
    code_col: str
    name_col: str
    manager_code_col: str | None
    manager_name_col: str | None
    date_col: str
    first_month_col: str
    active_col: str | None
    manpower_start: str
    active_rate_start: str
    premium_start: str
    qualified_start: str
    maintain_start: str
    award_start: str
    final_award_start: str | None
    total_col: str
    status_col: str
    supervisor_count_start: str | None = None
    org_award_start: str | None = None
    star_award_start: str | None = None


LAYOUTS = [
    SectionLayout(
        sheet="入职主管",
        section="引才奖-主管",
        code_col="A",
        name_col="B",
        manager_code_col="C",
        manager_name_col="D",
        date_col="E",
        first_month_col="F",
        active_col="G",
        manpower_start="H",
        active_rate_start="N",
        premium_start="T",
        qualified_start="Z",
        maintain_start="AF",
        award_start="AL",
        final_award_start="AS",
        total_col="AY",
        status_col="AZ",
    ),
    SectionLayout(
        sheet="入职经理",
        section="引才奖-经理",
        code_col="A",
        name_col="B",
        manager_code_col=None,
        manager_name_col=None,
        date_col="C",
        first_month_col="G",
        active_col=None,
        manpower_start="H",
        supervisor_count_start="N",
        active_rate_start="T",
        premium_start="Z",
        qualified_start="AF",
        maintain_start="AL",
        award_start="AR",
        final_award_start=None,
        total_col="AX",
        status_col="AZ",
    ),
    SectionLayout(
        sheet="晋升主管",
        section="晋升育成",
        code_col="A",
        name_col="B",
        manager_code_col="C",
        manager_name_col="D",
        date_col="E",
        first_month_col="F",
        active_col="G",
        manpower_start="H",
        active_rate_start="N",
        premium_start="T",
        qualified_start="Z",
        maintain_start="AF",
        award_start="AL",
        final_award_start="BD",
        total_col="BJ",
        status_col="BK",
        org_award_start="AR",
        star_award_start="AX",
    ),
]


def _col(col: str) -> int:
    return column_index_from_string(col)


def _offset_col(start_col: str, month: int) -> int:
    return _col(start_col) + MONTHS.index(month)


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _number(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _int_or_none(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _flag(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = _text(value)
    return text in {"1", "是", "达标", "维持", "Y", "y", "true", "True"}


def _date_iso(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return text


def _entry_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _get(ws, row: int, col: str) -> Any:
    return ws.cell(row, _col(col)).value


def _build_row(ws, row: int, layout: SectionLayout) -> dict | None:
    code = _text(_get(ws, row, layout.code_col))
    name = _text(_get(ws, row, layout.name_col))
    if not code and not name:
        return None

    monthly = []
    for month in MONTHS:
        award = _number(ws.cell(row, _offset_col(layout.award_start, month)).value)
        final_award = (
            _number(ws.cell(row, _offset_col(layout.final_award_start, month)).value)
            if layout.final_award_start
            else award
        )
        org_award = (
            _number(ws.cell(row, _offset_col(layout.org_award_start, month)).value)
            if layout.org_award_start
            else 0.0
        )
        star_award = (
            _number(ws.cell(row, _offset_col(layout.star_award_start, month)).value)
            if layout.star_award_start
            else 0.0
        )
        monthly.append(
            {
                "month": month,
                "schemeMonth": max(0, month - int(_number(_get(ws, row, layout.first_month_col), month)) + 1),
                "manpower": _number(ws.cell(row, _offset_col(layout.manpower_start, month)).value),
                "supervisorCount": (
                    _number(ws.cell(row, _offset_col(layout.supervisor_count_start, month)).value)
                    if layout.supervisor_count_start
                    else None
                ),
                "activeRate": _number(ws.cell(row, _offset_col(layout.active_rate_start, month)).value),
                "standardPremium": _number(ws.cell(row, _offset_col(layout.premium_start, month)).value),
                "qualified": _flag(ws.cell(row, _offset_col(layout.qualified_start, month)).value),
                "maintained": _flag(ws.cell(row, _offset_col(layout.maintain_start, month)).value),
                "award": award,
                "organizationAward": org_award,
                "starAward": star_award,
                "finalAward": final_award,
            }
        )

    return {
        "section": layout.section,
        "teamCode": code,
        "teamName": name,
        "managerCode": _text(_get(ws, row, layout.manager_code_col)) if layout.manager_code_col else "",
        "managerName": _text(_get(ws, row, layout.manager_name_col)) if layout.manager_name_col else "",
        "entryOrPromotionDate": _date_iso(_get(ws, row, layout.date_col)),
        "firstMonth": _int_or_none(_get(ws, row, layout.first_month_col)),
        "activeSupervisor": _flag(_get(ws, row, layout.active_col)) if layout.active_col else None,
        "rowTotalAward": _number(_get(ws, row, layout.total_col)),
        "status": _text(_get(ws, row, layout.status_col)),
        "monthly": monthly,
    }


def _parse_sections(workbook) -> list[dict]:
    rows: list[dict] = []
    for layout in LAYOUTS:
        if layout.sheet not in workbook.sheetnames:
            continue
        ws = workbook[layout.sheet]
        for row_idx in range(3, ws.max_row + 1):
            item = _build_row(ws, row_idx, layout)
            if item:
                rows.append(item)
    return rows


def _section_summary(rows: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = {}
    for row in rows:
        item = grouped.setdefault(
            row["section"],
            {
                "section": row["section"],
                "teamCount": 0,
                "qualifiedTeamCount": 0,
                "maintainedTeamCount": 0,
                "eliminatedTeamCount": 0,
                "totalAward": 0.0,
            },
        )
        item["teamCount"] += 1
        if any(month["qualified"] for month in row["monthly"]):
            item["qualifiedTeamCount"] += 1
        if "淘汰" in row.get("status", ""):
            item["eliminatedTeamCount"] += 1
        elif "维持" in row.get("status", ""):
            item["maintainedTeamCount"] += 1
        item["totalAward"] += sum(month["finalAward"] for month in row["monthly"])
    return list(grouped.values())


def _build_summary(rows: list[dict]) -> dict:
    latest_month = None
    total_award = 0.0
    recruit_award = 0.0
    organization_award = 0.0
    star_award = 0.0
    qualified_teams = 0
    maintained_teams = 0
    eliminated_teams = 0

    for row in rows:
        qualified = any(month["qualified"] for month in row["monthly"])
        if qualified:
            qualified_teams += 1
        if "淘汰" in row.get("status", ""):
            eliminated_teams += 1
        elif "维持" in row.get("status", ""):
            maintained_teams += 1
        for month in row["monthly"]:
            if any(_number(month[key]) for key in ("award", "organizationAward", "starAward", "finalAward")):
                latest_month = max(latest_month or month["month"], month["month"])
            total_award += _number(month["finalAward"])
            if row["section"] in {"引才奖-主管", "引才奖-经理"}:
                recruit_award += _number(month["finalAward"])
            organization_award += _number(month["organizationAward"])
            star_award += _number(month["starAward"])

    return {
        "totalTeams": len(rows),
        "qualifiedTeams": qualified_teams,
        "maintainedTeams": maintained_teams,
        "eliminatedTeams": eliminated_teams,
        "totalAward": round(total_award, 2),
        "recruitAward": round(recruit_award, 2),
        "organizationAward": round(organization_award, 2),
        "starAward": round(star_award, 2),
        "latestMonth": latest_month,
        "sections": _section_summary(rows),
    }


def _collect_warnings(workbook, formula_workbook) -> list[dict]:
    warnings: list[dict] = []
    missing = [layout.sheet for layout in LAYOUTS if layout.sheet not in workbook.sheetnames]
    if missing:
        warnings.append(
            {
                "level": "high",
                "title": "缺少测算工作表",
                "message": f"当前文件缺少 {', '.join(missing)}，对应测算结果不会展示。",
            }
        )

    if "参数表" in workbook.sheetnames:
        start_value = workbook["参数表"]["B2"].value
        start_date = _entry_date_value(start_value)
        if start_date != ENTRY_START:
            warnings.append(
                {
                    "level": "high",
                    "title": "方案统计人力起算日需修正",
                    "message": f"参数表 B2 当前为 {_date_iso(start_value) or start_value}，政策口径应为 {ENTRY_START.isoformat()}。",
                }
            )
    else:
        warnings.append({"level": "high", "title": "缺少参数表", "message": "无法核对方案期、维持系数和奖励档位。"})

    if "入职主管" in formula_workbook.sheetnames:
        ws = formula_workbook["入职主管"]
        manager_formula_defects = []
        for row_idx in range(3, ws.max_row + 1):
            formula = _text(ws.cell(row_idx, _col("AR")).value)
            if "VLOOKUP" in formula and "入职经理!$A:$F,7" in formula:
                manager_formula_defects.append(row_idx)
        if manager_formula_defects:
            warnings.append(
                {
                    "level": "medium",
                    "title": "入职主管页经理参与公式疑似失效",
                    "message": "AR列在 A:F 区间取第7列，超出查找范围，当前底稿可能将经理参与方案默认压为0。",
                }
            )

    field_sheets = {"业绩清单", "人力清单", "入职主管", "入职经理", "晋升主管"}
    header_text = " ".join(
        _text(cell.value)
        for sheet in workbook.worksheets
        if sheet.title in field_sheets
        for row in sheet.iter_rows(min_row=1, max_row=2)
        for cell in row
        if cell.value is not None
    )
    if "推荐人" not in header_text:
        warnings.append(
            {
                "level": "medium",
                "title": "推荐人奖励缺少独立字段",
                "message": "正式政策要求校验推荐人当月活动人力并单独计算1000元/月/团队，当前底稿未提供推荐人字段。",
            }
        )
    if not all(keyword in header_text for keyword in ["回访", "犹豫", "互保"]):
        warnings.append(
            {
                "level": "medium",
                "title": "有效保单条件尚不能完整校验",
                "message": "45日撤保退保、回执回访、犹豫期、自保互保等政策条件需补充源字段后才能自动判断。",
            }
        )

    missing_cache_cells = []
    for layout in LAYOUTS:
        if layout.sheet not in workbook.sheetnames or layout.sheet not in formula_workbook.sheetnames:
            continue
        cached_ws = workbook[layout.sheet]
        formula_ws = formula_workbook[layout.sheet]
        output_starts = [layout.qualified_start, layout.maintain_start, layout.award_start]
        output_starts.extend(
            value for value in [layout.final_award_start, layout.org_award_start, layout.star_award_start] if value
        )
        for row_idx in range(3, formula_ws.max_row + 1):
            if not (_text(_get(formula_ws, row_idx, layout.code_col)) or _text(_get(formula_ws, row_idx, layout.name_col))):
                continue
            output_cells = [formula_ws.cell(row_idx, _offset_col(start, month)) for start in output_starts for month in MONTHS]
            output_cells.extend(
                [formula_ws.cell(row_idx, _col(layout.total_col)), formula_ws.cell(row_idx, _col(layout.status_col))]
            )
            for formula_cell in output_cells:
                value = formula_cell.value
                if not (formula_cell.data_type == "f" or (isinstance(value, str) and value.startswith("="))):
                    continue
                if cached_ws.cell(formula_cell.row, formula_cell.column).value in (None, ""):
                    missing_cache_cells.append(f"{layout.sheet}!{formula_cell.coordinate}")
    if missing_cache_cells:
        preview = ", ".join(missing_cache_cells[:10])
        suffix = "等" if len(missing_cache_cells) > 10 else ""
        warnings.append(
            {
                "level": "high",
                "title": "关键公式缓存缺失",
                "message": f"{preview}{suffix} 未保存可读取的公式结果，请先用 Excel 完成重算并保存。",
            }
        )
    return warnings


def calculate_2026_org_dev_workbook(content: bytes, file_name: str) -> dict:
    workbook = load_workbook(BytesIO(content), data_only=True)
    formula_workbook = load_workbook(BytesIO(content), data_only=False)
    rows = _parse_sections(workbook)
    warnings = _collect_warnings(workbook, formula_workbook)
    if not rows:
        warnings.append(
            {
                "level": "high",
                "title": "测算明细为空",
                "message": "三个方案工作表均未解析到有效人员行，不能生成成功批次。",
            }
        )
    summary = _build_summary(rows)
    source_sheets = [{"name": name, "rows": workbook[name].max_row, "columns": workbook[name].max_column} for name in workbook.sheetnames]

    return {
        "scheme": {
            "id": SCHEME_ID,
            "name": SCHEME_NAME,
            "ruleVersion": RULE_VERSION,
        },
        "summary": summary,
        "details": {"rows": rows},
        "warnings": warnings,
        "definitions": RULE_DEFINITIONS,
        "sourceAudit": {
            "fileName": file_name,
            "sheetCount": len(workbook.sheetnames),
            "sheets": source_sheets,
            "calculationBasis": "正式PDF为主、Excel底稿为当前测算展示依据。",
        },
    }
