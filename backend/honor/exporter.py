"""Excel export for honor alliance results."""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .repository import fetch_summary, fetch_table

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="1F4E78", bold=True, size=13)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def build_honor_export_workbook(batch_id: int) -> bytes:
    summary = fetch_summary(batch_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "总览"
    _write_kv(ws, "星钻联盟荣誉体系总览", summary.get("overview", {}))

    sheets = [
        ("机构汇总", "honor_org_summary"),
        ("人员汇总", "honor_person_summary"),
        ("月度明细", "honor_person_month"),
        ("季度奖励测算", "honor_quarter_rewards"),
        ("异常清单", "honor_exceptions"),
        ("字段审计", "honor_field_audit_results"),
    ]
    for title, table in sheets:
        _write_table(wb.create_sheet(title), fetch_table(table, batch_id, limit=5000))

    rule_ws = wb.create_sheet("规则口径")
    _write_rows(
        rule_ws,
        ["规则", "口径"],
        [
            ["会员等级", "按当前累计钻石匹配最高门槛；低于3颗为未入会。"],
            ["OTO月度达标", "月度折算保费>=20000元且长险>=1件，获1颗钻石。"],
            ["证保月度达标", "月度折算保费>=30000元且长险>=1件，获1颗钻石。"],
            ["证保保号", "证保当月有长险件但未达标，不扣减。"],
            ["月末非在职", "月末非在职人员当前钻石清零。"],
            ["奖励", "本表为预计/测算奖励，不代表最终发放金额。"],
        ],
    )
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _write_kv(ws, title: str, payload: dict[str, Any]) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    row = 3
    for key, value in payload.items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1
    _style_sheet(ws)


def _write_table(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["暂无数据"])
        _style_sheet(ws)
        return
    headers = list(rows[0].keys())
    _write_rows(ws, headers, [[row.get(h) for h in headers] for row in rows])


def _write_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    _style_sheet(ws)


def _style_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
    for idx, col in enumerate(ws.columns, start=1):
        max_len = 10
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 36)
    ws.freeze_panes = "A2"

