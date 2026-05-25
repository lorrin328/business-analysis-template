from io import BytesIO
import os
import sys

import pytest

pytest.importorskip("fastapi")
pytest.importorskip("openpyxl")

from fastapi.testclient import TestClient
from openpyxl import load_workbook

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.join(ROOT, "backend"))

from main import app


client = TestClient(app)


def test_excel_export_returns_workbook_with_module_sheets():
    resp = client.get("/api/export/excel?year=2026")

    assert resp.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers["content-type"]
    assert "filename*=UTF-8''" in resp.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp.content), data_only=True)
    expected_sheets = [
        "导出说明",
        "KPI概览",
        "目标设置",
        "机构维度",
        "平台趋势",
        "产品结构",
        "产品明细",
        "交期结构",
        "队伍分析",
        "产品参数",
    ]
    assert wb.sheetnames == expected_sheets
    assert wb["KPI概览"]["A2"].value == "指标"
    assert wb["KPI概览"]["A3"].value == "期交保费达成率"
    assert wb["目标设置"]["A2"].value == "年份"
    assert wb["机构维度"]["A2"].value == "机构"


def test_excel_export_applies_table_presentation():
    resp = client.get("/api/export/excel?year=2026")
    wb = load_workbook(BytesIO(resp.content), data_only=True)

    ws = wb["KPI概览"]
    assert ws.freeze_panes == "A3"
    assert ws.auto_filter.ref.startswith("A2:")
    assert ws["A1"].font.bold is True
    assert ws["A2"].font.bold is True
    assert ws["A2"].fill.fgColor.rgb in {"001F4E78", "FF1F4E78"}
