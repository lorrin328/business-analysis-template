from io import BytesIO
from datetime import datetime

import pytest

pytest.importorskip("fastapi")
from fastapi.testclient import TestClient
from openpyxl import Workbook

from main import app


def _login(client, username="admin", password="Test-only-admin-2026!"):
    resp = client.post("/api/auth/login", json={"username": username, "password": password})
    assert resp.status_code == 200
    return resp.json()["data"]


def _headers(token):
    return {"Authorization": f"Bearer {token}"}


def _set_month_values(ws, row, start_col, values):
    from openpyxl.utils import column_index_from_string

    base = column_index_from_string(start_col)
    for offset, value in enumerate(values):
        ws.cell(row, base + offset).value = value


def _build_scheme_workbook():
    wb = Workbook()
    params = wb.active
    params.title = "参数表"
    params["B1"] = datetime(2026, 7, 31)
    params["B2"] = datetime(2026, 7, 1)
    params["B3"] = 0.7
    params["B4"] = 0.02

    supervisor = wb.create_sheet("入职主管")
    supervisor.append(["主管工号", "主管姓名", "经理工号", "经理姓名", "入司日期", "方案首月", "在职主管"])
    supervisor.append(["主管工号", "主管姓名", "经理工号", "经理姓名", "入司日期", "方案首月", "在职主管"])
    supervisor.append(["S001", "主管甲", "M001", "经理甲", datetime(2026, 7, 1), 7, 1])
    _set_month_values(supervisor, 3, "H", [3, 0, 0, 0, 0, 0])
    _set_month_values(supervisor, 3, "N", [0.5, 0, 0, 0, 0, 0])
    _set_month_values(supervisor, 3, "T", [4, 0, 0, 0, 0, 0])
    _set_month_values(supervisor, 3, "Z", [1, 0, 0, 0, 0, 0])
    _set_month_values(supervisor, 3, "AF", [1, 0, 0, 0, 0, 0])
    _set_month_values(supervisor, 3, "AL", [2000, 0, 0, 0, 0, 0])
    _set_month_values(supervisor, 3, "AS", [2000, 0, 0, 0, 0, 0])
    supervisor["AY3"] = 2000
    supervisor["AZ3"] = "维持"

    manager = wb.create_sheet("入职经理")
    manager.append(["经理工号", "经理姓名", "入司时间", "业务模式名称", "销售机构名称", "月末在职人力", "方案首月"])
    manager.append(["经理工号", "经理姓名", "入司时间", "业务模式名称", "销售机构名称", "月末在职人力", "方案首月"])
    manager.append(["M001", "经理甲", datetime(2026, 7, 1), "OTO", "上海", 1, 7])
    _set_month_values(manager, 3, "H", [4, 0, 0, 0, 0, 0])
    _set_month_values(manager, 3, "N", [2, 0, 0, 0, 0, 0])
    _set_month_values(manager, 3, "T", [0.5, 0, 0, 0, 0, 0])
    _set_month_values(manager, 3, "Z", [10, 0, 0, 0, 0, 0])
    _set_month_values(manager, 3, "AF", [1, 0, 0, 0, 0, 0])
    _set_month_values(manager, 3, "AL", [1, 0, 0, 0, 0, 0])
    _set_month_values(manager, 3, "AR", [3000, 0, 0, 0, 0, 0])
    manager["AX3"] = 3000
    manager["AZ3"] = "维持"

    promotion = wb.create_sheet("晋升主管")
    promotion.append(["主管工号", "主管姓名", "经理工号", "经理姓名", "晋升日期", "方案首月", "在职主管"])
    promotion.append(["主管工号", "主管姓名", "经理工号", "经理姓名", "晋升日期", "方案首月", "在职主管"])
    promotion.append(["P001", "晋升甲", "M001", "经理甲", datetime(2026, 7, 1), 7, 1])
    _set_month_values(promotion, 3, "H", [3, 0, 0, 0, 0, 0])
    _set_month_values(promotion, 3, "N", [0.5, 0, 0, 0, 0, 0])
    _set_month_values(promotion, 3, "T", [4, 0, 0, 0, 0, 0])
    _set_month_values(promotion, 3, "Z", [1, 0, 0, 0, 0, 0])
    _set_month_values(promotion, 3, "AF", [1, 0, 0, 0, 0, 0])
    _set_month_values(promotion, 3, "AL", [2000, 0, 0, 0, 0, 0])
    _set_month_values(promotion, 3, "AR", [2000, 0, 0, 0, 0, 0])
    _set_month_values(promotion, 3, "AX", [80, 0, 0, 0, 0, 0])
    _set_month_values(promotion, 3, "BD", [4080, 0, 0, 0, 0, 0])
    promotion["BJ3"] = 4080
    promotion["BK3"] = "维持"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def test_scheme_options_and_dedicated_upload(auth_db):
    client = TestClient(app)
    admin = _login(client)
    headers = _headers(admin["token"])

    options = client.get("/api/scheme/options", headers=headers)
    assert options.status_code == 200
    assert options.json()["data"]["defaultSchemeId"] == "2026-org-dev-policy"
    assert options.json()["data"]["schemes"][0]["name"] == "2026年组发政策"

    content = _build_scheme_workbook()
    uploaded = client.post(
        "/api/scheme/upload",
        headers=headers,
        data={"schemeId": "2026-org-dev-policy"},
        files={"tracking": ("组织发展追踪模板.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert uploaded.status_code == 200
    data = uploaded.json()["data"]
    assert data["batch"]["schemeName"] == "2026年组发政策"
    assert data["summary"]["totalTeams"] == 3
    assert data["summary"]["qualifiedTeams"] == 3
    assert data["summary"]["totalAward"] == 9080
    assert data["summary"]["organizationAward"] == 2000
    assert data["summary"]["starAward"] == 80
    assert any(item["title"] == "推荐人奖励缺少独立字段" for item in data["warnings"])

    latest = client.get("/api/scheme/latest?schemeId=2026-org-dev-policy", headers=headers)
    assert latest.status_code == 200
    assert latest.json()["data"]["batch"]["id"] == data["batch"]["id"]


def test_scheme_upload_is_separate_permission(auth_db):
    client = TestClient(app)
    registered = client.post("/api/auth/register", json={"username": "scheme_normal", "password": "normal-pass-123"})
    assert registered.status_code == 200
    token = registered.json()["data"]["token"]
    headers = _headers(token)

    assert client.get("/api/scheme/options", headers=headers).status_code == 200
    blocked = client.post(
        "/api/scheme/upload",
        headers=headers,
        data={"schemeId": "2026-org-dev-policy"},
        files={"tracking": ("组织发展追踪模板.xlsx", _build_scheme_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert blocked.status_code == 403


def test_scheme_upload_rejects_missing_sheets_without_replacing_latest(auth_db):
    client = TestClient(app)
    admin = _login(client)
    headers = _headers(admin["token"])
    valid = client.post(
        "/api/scheme/upload",
        headers=headers,
        data={"schemeId": "2026-org-dev-policy"},
        files={"tracking": ("valid.xlsx", _build_scheme_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    valid_batch_id = valid.json()["data"]["batch"]["id"]

    wb = Workbook()
    wb.active.title = "无关工作表"
    bio = BytesIO()
    wb.save(bio)
    rejected = client.post(
        "/api/scheme/upload",
        headers=headers,
        data={"schemeId": "2026-org-dev-policy"},
        files={"tracking": ("invalid.xlsx", bio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    latest = client.get("/api/scheme/latest?schemeId=2026-org-dev-policy", headers=headers)

    assert rejected.status_code == 422
    assert latest.json()["data"]["batch"]["id"] == valid_batch_id
    titles = [item["title"] for item in rejected.json()["detail"]["warnings"]]
    assert "缺少测算工作表" in titles
    assert "测算明细为空" in titles
