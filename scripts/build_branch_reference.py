"""从证保报表参数表生成不含人员字段的网点参考表。"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FIELDS = [
    "参考编号",
    "网点类型",
    "证券网点",
    "归属主体",
    "所在省",
    "所在市",
    "网点等级",
    "机构类项目",
    "机构类项目细分",
    "本地异地",
    "纳入常规网点数",
    "源表行号",
]


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def infer_parent(branch: str, project: str, branch_type: str) -> str:
    if branch_type == "转介绍网点":
        return "广发证券股份有限公司"
    if "广发" in branch or "广发" in project:
        return "广发证券股份有限公司"
    if "银河" in branch or "银河" in project:
        return "中国银河证券股份有限公司"
    if "中信" in branch or "中信" in project:
        return "中信证券股份有限公司"
    return ""


def extract_rows(source: Path) -> list[dict[str, str]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if "参数表" not in workbook.sheetnames:
            raise ValueError(f"{source.name} 不含“参数表”工作表")
        sheet = workbook["参数表"]
        if clean_text(sheet["AA1"].value) != "证券网点":
            raise ValueError(f"{source.name} 参数表AA1不是“证券网点”")
        result: list[dict[str, str]] = []
        ranges = [
            ("常规网点", range(2, 149), "是", "REG"),
            ("转介绍网点", range(151, 238), "否", "REF"),
        ]
        for branch_type, row_numbers, included, prefix in ranges:
            index = 0
            for row_number in row_numbers:
                branch = clean_text(sheet.cell(row_number, 27).value)
                if not branch:
                    continue
                index += 1
                project = clean_text(sheet.cell(row_number, 34).value)
                result.append(
                    {
                        "参考编号": f"{prefix}-{index:03d}",
                        "网点类型": branch_type,
                        "证券网点": branch,
                        "归属主体": infer_parent(branch, project, branch_type),
                        "所在省": clean_text(sheet.cell(row_number, 28).value),
                        "所在市": clean_text(sheet.cell(row_number, 29).value),
                        "网点等级": clean_text(sheet.cell(row_number, 33).value),
                        "机构类项目": project,
                        "机构类项目细分": clean_text(sheet.cell(row_number, 35).value),
                        "本地异地": clean_text(sheet.cell(row_number, 36).value),
                        "纳入常规网点数": included,
                        "源表行号": str(row_number),
                    }
                )
        return result
    finally:
        workbook.close()


def write_reference(rows: list[dict[str, str]], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="生成项目内证保网点标准参考表")
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    if not args.source.is_file():
        raise FileNotFoundError(args.source)
    rows = extract_rows(args.source)
    regular = sum(row["网点类型"] == "常规网点" for row in rows)
    referral = sum(row["网点类型"] == "转介绍网点" for row in rows)
    if regular != 147 or referral != 86:
        raise ValueError(f"网点数量与已核验口径不符：常规{regular}，转介绍{referral}")
    if len({row["证券网点"] for row in rows}) != len(rows):
        raise ValueError("参考表存在重复证券网点名称")
    write_reference(rows, args.output)
    print(f"已生成 {args.output.resolve()}：常规{regular}个，转介绍{referral}个")


if __name__ == "__main__":
    main()
